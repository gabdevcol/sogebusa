import openpyxl


def ProcesarListaDeDescarga(book):
    lista_de_descarga = {
        "head": {"vessel": str(), "imo_number": int(), "nationality_of_ship": str(), "vayage": str(), "eta": str(),
                 "captains_name": str(), "ship_type": str()},
        "body": [{"linea": str(), "shipper": str(), "consignee": str(), "container_no": str(), "container": int(),
                  "type": str(), "owner": str(), "soc": str(), "package_type": str(), "imo": str(), "temp": str(),
                  "1_seal": str(), "2_seal": str(), "bls": str(), "description": str(), "gross": float(),
                  "puerto_origen": str(), "et": str(), "puerto_transbordo": str(), "puerto_descarga": str(),
                  "sidun": str()}
                 ]
    }

    try:

        hoja = book["FINAL"]

        for fila in hoja.values:
            print(fila)

        return lista_de_descarga

    except Exception as error:
        print(error)


def ProcesarListaDeEmbarque(book):
    lista_de_embarque = {
        "import": {"head": {"titulo": str(), "buque": str(), "viaje": str(), "puerto_descarga": str()},
                   "body": [{"nbr": int(), "container_no": str(), "alm_patio": str(), "sello1": str(),
                             "sello2": str(), "type": str(), "pol": str(), "pod": str(), "cat": str(),
                             "imo": int(), "un_#": int(), "temp": float(), "o_h": float(),
                             "o_l": float(),
                             "o_r": float(), "wt_mt": float(), "wt_kg": int(), "sts": str(),
                             "line": str(), "vsl_slot": int(), "remark": str()}]},
        "export": {"head": {"titulo": str(), "buque": str(), "viaje": str(), "puerto_descarga": str()},
                   "body": [{"nbr": int(), "container_no": str(), "sello1": str(), "sello2": str(),
                             "type": str(), "pol": str(), "pod": str(), "cat": str(), "imo": int(),
                             "un_#": int(), "temp": float(), "o_h": float(), "o_l": float(),
                             "o_r": float(),
                             "wt_mt": float(), "wt_kg": int(), "sts": str(), "line": str(),
                             "vsl_slot": int(), "remark": str()}]},
        "restow": {"head": {"titulo": str(), "buque": str(), "viaje": str(), "puerto_descarga": str()},
                   "body": [{"nbr": int(), "container_no": str(), "type": str(), "pol": str(), "pod": str(),
                             "cat": str(), "imo": int(), "un_#": int(), "previous_position": int(),
                             "new_position": int(), "restow_account": str(), "temp": float(),
                             "o_h": float(),
                             "o_l": float(), "o_r": float(), "wt_mt": float(), "wt_kg": int(),
                             "sts": str(), "line": str(), "remark": str()}]},
        "cancel": {"head": {"titulo": str(), "buque": str(), "viaje": str(), "puerto_descarga": str()},
                   "body": [{"nbr": int(), "container_no": str(), "type": str(), "pol": str(), "pod": str(),
                             "cat": str(), "imo": int(), "un_#": int(), "temp": float(),
                             "o_h": float(),
                             "o_l": float(), "o_r": float(), "wt_mt": float(), "wt_kg": int(),
                             "sts": str(), "line": str(), "remark": str()
                             }]}
    }

    try:

        for nom_hoja in book.sheetnames:
            hoja = book[nom_hoja]
            filas = hoja.max_row
            columnas = hoja.max_column

            lista_de_embarque[nom_hoja.lower()]["body"] *= (filas - 2)
            lista_de_embarque[nom_hoja.lower()]["head"]["titulo"] = hoja.cell(1, 1).value

            for fila in range(3, filas + 1):
                indice_lista_hoja = fila - 3
                nuevo_registro = {}

                for col in range(1, columnas + 1):
                    clave = list(lista_de_embarque[nom_hoja.lower()]["body"][indice_lista_hoja])[col - 1]
                    nuevo_registro[clave] = hoja.cell(fila, col).value

                lista_de_embarque[nom_hoja.lower()]["body"][indice_lista_hoja] = nuevo_registro

        return lista_de_embarque

    except Exception as error:
        print(error)


def ProcesarListaDeEquiposPrioridad(book):
    lista_de_equipos_prioridad = {
        "head": {"titulo": str()},
        "body": [{"id": int(), "contenedor": str(), "tipo": str(), "buque": str(), "viaje": str(), "arribo": str(),
                  "pod": str()}]
    }

    try:

        hoja = book[book.sheetnames[0]]

        for fila in hoja.values:
            if type(fila[0]) == int:
                lista_de_equipos_prioridad["body"].append({"id": fila[0], "contenedor": fila[1], "tipo": fila[2],
                                                           "buque": fila[3], "viaje": fila[4], "arribo": fila[5],
                                                           "pod": fila[6]})

        lista_de_equipos_prioridad["body"].pop(0)
        return lista_de_equipos_prioridad

    except Exception as error:
        print(error)


def ProcesarTabuladorParaEquiposInoperativos(book):
    tabulador_para_equipos_inoperativos = {
        "head": {"titulo": str(), "fecha": str()},
        "body": [{"codigo": str(), "parte_cont": str(), "contenido": str(), "hh_para_reparacion": float(),
                  "costo_hh_para_reparacion": float(), "costo_materiales_proveedor": float(), "costo_ronald_foguera": float(),
                  "costo_sogebusa": float(), "diferencia_en_factura": str(), "hh_para_reparacion-2": float(),
                  "costo_hh_para_reparacion-2": float(), "costo_materiales_equipos": float()}]
    }
    try:

        hoja = book[book.sheetnames[0]]
        filtrador = 0

        for fila in hoja.values:
            if filtrador > 0:
                tabulador_para_equipos_inoperativos["body"].append(
                    {"codigo": fila[1], "parte_cont": fila[2], "contenido": fila[3], "hh_para_reparacion": fila[4],
                     "costo_hh_para_reparacion": fila[5], "costo_materiales_proveedor": fila[6],
                     "costo_ronald_foguera": fila[7], "costo_sogebusa": fila[8], "diferencia_en_factura": fila[9],
                     "hh_para_reparacion-2": fila[10], "costo_hh_para_reparacion-2": fila[11],
                     "costo_materiales_equipos": fila[12]})

            elif fila[1] == "CODIGO":
                filtrador += 1

        tabulador_para_equipos_inoperativos["body"].pop(0)
        return tabulador_para_equipos_inoperativos

    except Exception as error:
        print(error)


def ProcesarReporteDeAcarreos(book):
    reporte_de_acarreos = {
        "head": {},
        "body": [{"item": int(), "siglas_numeros": str(), "tipo": str(), "almacen": str(), "linea": str(),
                  "get_out_export": str(), "eir_2": int(), "consigned": str(), "pod": str(), "pod_2": str()}]
    }

    try:

        hoja = book[book.sheetnames[0]]

        for fila in hoja.values:
            if type(fila[0]) == int:
                reporte_de_acarreos["body"].append({"item": fila[0], "siglas_numeros": fila[1], "tipo": fila[2],
                                                    "almacen": fila[3], "linea": fila[4], "get_out_export": fila[5],
                                                    "eir_2": fila[6], "consigned": fila[7], "pod": fila[8],
                                                    "pod_2": fila[9]})

        reporte_de_acarreos["body"].pop(0)
        return reporte_de_acarreos

    except Exception as error:
        print(error)


def DocumentoProcesar(doc):
    archivo_ruta = doc

    try:
        book = openpyxl.load_workbook(filename=archivo_ruta, data_only=True)
        hojas = book.sheetnames
        # print(hojas)

        # if len(hojas) > 1:
        #     return ProcesarListaDeEmbarque(book)

        # return ProcesarReporteDeAcarreos(book)
        # return ProcesarListaDeEquiposPrioridad(book)
        return ProcesarTabuladorParaEquiposInoperativos(book)

    except Exception as error:
        print(error)


doc = "Altransaga/5 - TABULADOR DE REPARACIONES.xlsx"

print(DocumentoProcesar(doc))
